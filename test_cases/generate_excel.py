#!/usr/bin/env python3
"""
Generate Excel spreadsheet with functional test cases for EBEK OSCE Platform
Based on actual codebase analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os

# Import test cases
from functional_test_cases import functional_test_cases

def create_test_cases_worksheet(ws, test_cases):
    """Create Test Cases worksheet"""
    headers = ["Test ID", "Category", "Priority", "Test Name", "Description",
               "Preconditions", "Test Steps", "Test Data", "Expected Results",
               "Actual Results", "Status", "Notes", "Related Files"]

    # Style header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Add test data
    for row_num, tc in enumerate(test_cases, 2):
        ws.cell(row=row_num, column=1).value = tc['id']
        ws.cell(row=row_num, column=2).value = tc['category']
        ws.cell(row=row_num, column=3).value = tc['priority']
        ws.cell(row=row_num, column=4).value = tc['test_name']
        ws.cell(row=row_num, column=5).value = tc['description']
        ws.cell(row=row_num, column=6).value = tc['preconditions']
        ws.cell(row=row_num, column=7).value = tc['steps']
        ws.cell(row=row_num, column=8).value = tc['test_data']
        ws.cell(row=row_num, column=9).value = tc['expected']
        ws.cell(row=row_num, column=10).value = tc['actual']
        ws.cell(row=row_num, column=11).value = tc['status']
        ws.cell(row=row_num, column=12).value = ""
        ws.cell(row=row_num, column=13).value = tc['file']

        # Apply formatting
        for col in range(1, 14):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            if col in [5, 6, 7, 9]:  # Wrap text for longer columns
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Color-code status
        status_cell = ws.cell(row=row_num, column=11)
        if tc['status'] == 'Not Run':
            status_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 25

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add filters
    ws.auto_filter.ref = f"A1:M{len(test_cases) + 1}"

def create_coverage_worksheet(ws, test_cases):
    """Create Coverage Matrix worksheet"""
    categories = sorted(set(tc['category'] for tc in test_cases))

    # Add title and summary
    ws['A1'] = "Test Coverage Matrix"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Add category counts
    ws['A3'] = "Category"
    ws['B3'] = "Test Count"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ['A3', 'B3']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font

    row = 4
    for category in categories:
        ws[f'A{row}'] = category
        count = len([tc for tc in test_cases if tc['category'] == category])
        ws[f'B{row}'] = count
        row += 1

    # Summary statistics
    row += 2
    ws[f'A{row}'] = "Priority Breakdown"
    ws[f'A{row}'].font = Font(bold=True, size=11)

    row += 1
    ws[f'A{row}'] = "Priority"
    ws[f'B{row}'] = "Count"
    for cell in [f'A{row}', f'B{row}']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font

    for priority in ['P0', 'P1', 'P2']:
        row += 1
        ws[f'A{row}'] = priority
        count = len([tc for tc in test_cases if tc['priority'] == priority])
        ws[f'B{row}'] = count

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

def create_execution_log_worksheet(ws):
    """Create Execution Log template worksheet"""
    headers = ["Date", "Time", "Tester Name", "Test ID", "Category", "Test Name",
               "Status", "Notes", "Evidence Link"]

    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Add sample row
    ws.cell(row=2, column=1).value = datetime.datetime.now().strftime("%Y-%m-%d")
    ws.cell(row=2, column=2).value = datetime.datetime.now().strftime("%H:%M")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 20

def create_summary_dashboard_worksheet(ws, test_cases):
    """Create Summary Dashboard"""
    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = "EBEK OSCE Platform - Functional Test Cases Dashboard"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20

    # Summary statistics
    row = 3
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=11)

    row = 4
    ws[f'A{row}'] = "Total Test Cases:"
    ws[f'B{row}'] = len(test_cases)
    ws[f'B{row}'].font = Font(bold=True)

    # By category
    row += 2
    ws[f'A{row}'] = "By Category:"
    ws[f'A{row}'].font = Font(bold=True)

    categories = {}
    for tc in test_cases:
        cat = tc['category']
        categories[cat] = categories.get(cat, 0) + 1

    for category in sorted(categories.keys()):
        row += 1
        ws[f'A{row}'] = f"  {category}"
        ws[f'B{row}'] = categories[category]

    # By priority
    row += 2
    ws[f'A{row}'] = "By Priority:"
    ws[f'A{row}'].font = Font(bold=True)

    priorities = {}
    for tc in test_cases:
        p = tc['priority']
        priorities[p] = priorities.get(p, 0) + 1

    for priority in sorted(priorities.keys()):
        row += 1
        ws[f'A{row}'] = f"  {priority}"
        ws[f'B{row}'] = priorities[priority]

    # Execution status
    row += 2
    ws[f'A{row}'] = "Test Execution Status"
    ws[f'A{row}'].font = Font(bold=True, size=11)

    row += 1
    ws[f'A{row}'] = "Not Run:"
    ws[f'B{row}'] = len([tc for tc in test_cases if tc['status'] == 'Not Run'])

    row += 1
    ws[f'A{row}'] = "Passed:"
    ws[f'B{row}'] = 0
    ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    row += 1
    ws[f'A{row}'] = "Failed:"
    ws[f'B{row}'] = 0
    ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

# Create workbook
print("🚀 Creating Excel workbook...")
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Create worksheets
print("📋 Creating Test Cases worksheet...")
ws_cases = wb.create_sheet("Test Cases", 0)
create_test_cases_worksheet(ws_cases, functional_test_cases)

print("📊 Creating Coverage Matrix worksheet...")
ws_coverage = wb.create_sheet("Coverage Matrix", 1)
create_coverage_worksheet(ws_coverage, functional_test_cases)

print("📝 Creating Execution Log worksheet...")
ws_log = wb.create_sheet("Execution Log", 2)
create_execution_log_worksheet(ws_log)

print("📈 Creating Summary Dashboard worksheet...")
ws_summary = wb.create_sheet("Summary Dashboard", 3)
create_summary_dashboard_worksheet(ws_summary, functional_test_cases)

# Save workbook
output_file = "/Users/sanujphilip/Desktop/Ebek/test_cases/EBEK_Functional_Test_Cases.xlsx"
wb.save(output_file)

print(f"\n✅ SUCCESS! Excel file created!")
print(f"📊 Total test cases: {len(functional_test_cases)}")
print(f"📁 File: {output_file}")
print(f"💾 File size: {round(os.path.getsize(output_file) / 1024, 2)} KB")
print("\n📋 Worksheets created:")
print("  1. Test Cases - 116 functional test cases with complete details")
print("  2. Coverage Matrix - Test count by category and priority")
print("  3. Execution Log - Template for tracking test execution")
print("  4. Summary Dashboard - Statistics and KPIs")
