import threading
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Group, Institution, Hospital, Learner, Assessor, SkillathonEvent, EbekUser, SchedularObject
from .onboarding_forms import (
    GroupForm, InstitutionForm, HospitalForm, LearnerForm,
    AssessorForm, SkillathonEventForm, BulkLearnerUploadForm
)
import csv
from django.http import HttpResponse, StreamingHttpResponse
from datetime import datetime
from django.contrib.auth import get_user_model
from assessments.utils_ses import send_email
import random
import string
from django.urls import reverse
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .onboarding_forms import LearnerForm
from .models import Learner, Institution, Hospital, SkillathonEvent
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from .firebase_sync import sync_user_to_firestore, sync_user_to_firebase_auth, batch_sync_users_to_firebase_auth, create_test_and_exam_assignments, DisableSignals, enable_all_signals
from django.db.models.signals import post_save, post_delete
from django.dispatch import Signal
from firebase_admin import firestore
import traceback
import json
import os
import uuid
from django.core.cache import cache
from .firebase_sync import (
    on_user_save, on_user_delete,
    on_institute_save, on_institution_delete,
    on_hospital_save, on_hospital_delete,
    on_learner_save, on_learner_delete,
    on_assessor_save, on_assessor_delete,
    on_skillathon_save, on_skillathon_delete,
    on_group_save, on_group_delete, batch_sync_users_to_firestore_with_progress
)
import threading
from django.db.models import Q
import logging

logger = logging.getLogger(__name__)

from django.conf import settings
import os
from dotenv import load_dotenv

load_dotenv()
firebase_database = os.getenv('FIREBASE_DATABASE')

db = firestore.client(database_id=firebase_database)

@login_required
def group_list(request):
    groups = Group.objects.all().order_by('-created_at')
    
    # Get search query
    search_query = request.GET.get('search', '').strip()
    
    # Get filter values as lists
    selected_types = request.GET.getlist('type')
    selected_statuses = request.GET.getlist('status')
    
    # Apply search filter
    if search_query:
        groups = groups.filter(
            Q(name__icontains=search_query) |
            Q(group_head__full_name__icontains=search_query) |
            Q(group_head__email__icontains=search_query)
        )
    
    # Apply filters
    if selected_types:
        groups = groups.filter(type__in=selected_types)
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        groups = groups.filter(is_active__in=active_status)
    
    paginator = Paginator(groups, 10)
    page = request.GET.get('page')
    groups = paginator.get_page(page)
    
    return render(request, 'assessments/onboarding/group_list.html', {
        'groups': groups,
        'selected_types': selected_types,
        'selected_statuses': selected_statuses,
        'search_query': search_query,
    })

@login_required
def group_create(request):
    User = get_user_model()
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            head_name = form.cleaned_data['group_head_name']
            head_email = form.cleaned_data['group_head_email']
            head_phone = form.cleaned_data['group_head_phone']
            group_name = form.cleaned_data['name'].strip()
            user = None
            
            # Check if group with same name already exists in Firebase
            existing_group = db.collection('Groups').where('name', '==', group_name).limit(1).stream()
            
            if list(existing_group):
                return JsonResponse({'error': 'A group with this name already exists'}, status=400)

            if head_name == '' or head_email == '' or head_phone == '':
                pass
            else:
                user, created = User.objects.get_or_create(
                    email=head_email,
                    defaults={
                        'is_active': True,
                    }
                )
                if created:
                    logger.info(f"Creating user {head_email}")
                    user.full_name = head_name
                    user.phone_number = head_phone
                    default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                    user.set_password(default_password)
                    user.save()
                    reset_link = request.build_absolute_uri(reverse('login'))
                    subject = 'Your Group Admin Account Created'
                    body = f"""Dear {head_name},\n\nYour group admin account has been created for the group {group_name}.\n\nUsername: {head_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                    send_email_thread = threading.Thread(target=send_email, args=(subject, body, [head_email]))
                    send_email_thread.start()
                else:
                    # Update name/phone if changed
                    updated = False
                    if user.full_name != head_name:
                        user.full_name = head_name
                        updated = True
                    if user.phone_number != head_phone:
                        user.phone_number = head_phone
                        updated = True
                    if updated:
                        user.save()
            group = form.save(commit=False)
            if user is not None:
                group.group_head = user
            if request.POST.get('is_active') == 'on':
                group.is_active = True
            else:
                group.is_active = False
            group.save()
            messages.success(request, 'Group created successfully.')
            return HttpResponse('OK')
    else:
        form = GroupForm()
    return render(request, 'assessments/onboarding/group_form.html', {'form': form, 'action': 'Create'})

@login_required
def group_edit(request, pk):
    User = get_user_model()
    group = get_object_or_404(Group, pk=pk)
    group_old_name = group.name
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            head_name = form.cleaned_data['group_head_name']
            head_email = form.cleaned_data['group_head_email']
            head_phone = form.cleaned_data['group_head_phone']
            group_name = form.cleaned_data['name']
            current_head = group.group_head
            
            # Check if group name is being changed and if it already exists in Firebase
            if group_old_name != group_name:
                existing_group = db.collection('Groups').where('name', '==', group_name).limit(1).stream()
                if list(existing_group):
                    return JsonResponse({'error': 'A group with this name already exists'}, status=400)
            
            if head_name != "" and head_email != "" and head_phone != "":
                if not current_head or current_head.email != head_email:
                    user, created = User.objects.get_or_create(
                        email=head_email,
                        defaults={
                            'is_active': True,
                        }
                    )
                    if created:
                        user.full_name = head_name
                        user.phone_number = head_phone
                        default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                        user.set_password(default_password)
                        user.save()
                        reset_link = request.build_absolute_uri(reverse('login'))
                        subject = 'Your Group Admin Account Created'
                        body = f"""Dear {head_name},\n\nYour group admin account has been created for the group {group_name}.\n\nUsername: {head_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                        send_email_thread = threading.Thread(target=send_email, args=(subject, body, [head_email]))
                        send_email_thread.start()
                    else:
                        updated = False
                        if user.full_name != head_name:
                            user.full_name = head_name
                            updated = True
                        if user.phone_number != head_phone:
                            user.phone_number = head_phone
                            updated = True
                        if updated:
                            user.save()
                    group.group_head = user
                else:
                    updated = False
                    if current_head.full_name != head_name:
                        current_head.full_name = head_name
                        updated = True
                    if current_head.phone_number != head_phone:
                        current_head.phone_number = head_phone
                        updated = True
                    if updated:
                        current_head.save()
            if request.POST.get('is_active') == 'on':
                group.is_active = True
            else:
                group.is_active = False
            group = form.save(commit=False)
            group.save()
            messages.success(request, 'Group updated successfully.')
            return HttpResponse('OK')
    else:
        form = GroupForm(instance=group, initial={
            'group_head_name': group.group_head.full_name if group.group_head else '',
            'group_head_email': group.group_head.email if group.group_head else '',
            'group_head_phone': group.group_head.phone_number if group.group_head else '',
        })
    return render(request, 'assessments/onboarding/group_form.html', {'form': form, 'action': 'Edit'})

@login_required
def group_delete(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        group.delete()
        messages.success(request, 'Group deleted successfully.')
        return redirect('group_list')
    return redirect('group_list')

# Institution Views
@login_required
def institution_list(request):
    if not (request.user.has_all_permissions() or 'view_institutes' in request.user.get_all_permissions()):
        return redirect('base')
    try:
        print("DEBUG: Syncing strength counts before loading institution list")
        sync_strength_counts(request)
    except Exception as e:
        print(f"DEBUG: Error syncing strength counts: {str(e)}")
    
    institutions = Institution.objects.all().order_by('-created_at') if request.user.has_all_permissions() else request.user.assigned_institutions.all().order_by('-created_at')
    print("HEREEEE")
    print(institutions)
    # Get all unique groups for the filter dropdown
    all_groups = Group.objects.filter(is_active=True, type="institution")
    
    # Get all unique states for the filter dropdown
    all_states = Institution.objects.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='') if request.user.has_all_permissions() else request.user.assigned_institutions.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='')
    
    # Filtering
    search_query = request.GET.get('query', '').strip()
    selected_groups = request.GET.getlist('group')
    print("DEBUG: Selected groups:", selected_groups)
    print(selected_groups)
    selected_states = request.GET.getlist('state')
    selected_statuses = request.GET.getlist('status')
    
    if search_query:
        institutions = institutions.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(district__icontains=search_query) |
            Q(state__icontains=search_query)
        )
    
    if selected_groups:
        institutions = institutions.filter(group_id__in=selected_groups)
    
    if selected_states:
        institutions = institutions.filter(state__in=selected_states)
    
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        institutions = institutions.filter(is_active__in=active_status)
    
    paginator = Paginator(institutions, 10)
    page = request.GET.get('page')
    institutions = paginator.get_page(page)

    print('DEBUG: Institutions:', institutions)
    
    return render(request, 'assessments/onboarding/institution_list.html', {
        'institutions': institutions,
        'all_groups': all_groups,
        'all_states': all_states,
        'selected_groups': selected_groups,
        'selected_states': selected_states,
        'selected_statuses': selected_statuses,
        'search_query': search_query,
    })

@login_required
def institution_create(request):
    if not (request.user.has_all_permissions() or 'create_institute' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    if request.method == 'POST':
        form = InstitutionForm(request.POST)
        if form.is_valid():
            institution_name = form.cleaned_data['name'].strip()
            onboarding_type = str(form.cleaned_data.get('onboarding_type', '')).lower()
            print(onboarding_type)
            user = None

            # Append " - B2C" suffix for B2C institutions
            if onboarding_type.lower() == "b2c":
                print("here")
                institution_name = str(institution_name) + " - B2C"

            # Validate uniqueness by onboarding type and name (case-insensitive)
            if Institution.objects.filter(name__iexact=institution_name, onboarding_type=onboarding_type).exists():
                return JsonResponse({'error': 'Institution already exists for this onboarding type.'}, status=400)

            # Create instance and modify name AFTER save(commit=False)
            institution = form.save(commit=False)

            # Set the modified name on the instance
            if onboarding_type.lower() == "b2c":
                institution.name = institution_name  # Use the modified name with " - B2C"

            if request.POST.get('is_active') == 'on':
                institution.is_active = True
            else:
                institution.is_active = False
            institution.save()
            messages.success(request, 'Institution created successfully.')
            return HttpResponse('OK')
    else:
        form = InstitutionForm()
    return render(request, 'assessments/onboarding/institution_form.html', {'form': form, 'action': 'Create'})

@login_required
def institution_edit(request, pk):
    if not (request.user.has_all_permissions() or 'edit_institute' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    institution = get_object_or_404(Institution, pk=pk)
    institution_old_name = institution.name
    if request.method == 'POST':
        form = InstitutionForm(request.POST, instance=institution)
        if form.is_valid():
            institution_name = form.cleaned_data['name'].strip()
            onboarding_type = str(form.cleaned_data.get('onboarding_type', '')).lower()
            # Enforce uniqueness within same onboarding type (exclude current)
            if Institution.objects.filter(name__iexact=institution_name, onboarding_type=onboarding_type).exclude(pk=pk).exists():
                return JsonResponse({'error': 'Institution already exists for this onboarding type.'}, status=400)
            
            if request.POST.get('is_active') == 'on':
                institution.is_active = True
            else:
                institution.is_active = False
            institution = form.save(commit=False)
            institution.save()
            messages.success(request, 'Institution updated successfully.')
            return HttpResponse('OK')
    else:
        form = InstitutionForm(instance=institution)
    return render(request, 'assessments/onboarding/institution_form.html', {'form': form, 'action': 'Edit'})

@login_required
def institution_delete(request, pk):
    if not (request.user.has_all_permissions() or 'delete_institute' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    institution = get_object_or_404(Institution, pk=pk)
    if request.method == 'POST':
        institution.delete()
        messages.success(request, 'Institution deleted successfully.')
        return redirect('institution_list')
    return redirect('institution_list')

@login_required
def institution_list_api(request):
    # Sync strength counts from Firebase before loading the data
    if not (request.user.has_all_permissions() or 'view_institutes' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    try:
        print("DEBUG: Syncing strength counts before loading institution API data")
        sync_strength_counts(request)
    except Exception as e:
        print(f"DEBUG: Error syncing strength counts: {str(e)}")
    
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()
    
    print(f"DEBUG: Institution search query received: '{search_query}'")
    
    institutions = Institution.objects.all().order_by('-created_at') if request.user.has_all_permissions() else request.user.assigned_institutions.all().order_by('-created_at')
    
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying institution search filter for: '{search_query}'")
        institutions = institutions.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(district__icontains=search_query) |
            Q(state__icontains=search_query)
        )
    print(request.GET)
    
    # Apply other filters
    selected_groups = request.GET.getlist('group')
    print(selected_groups)
    selected_states = request.GET.getlist('state')
    selected_statuses = request.GET.getlist('status')

    print(f"DEBUG: Group filters received: {selected_groups}")
    print(f"DEBUG: State filters received: {selected_states}")
    print(f"DEBUG: Status filters received: {selected_statuses}")

    if selected_groups:
        institutions = institutions.filter(group_id__in=selected_groups)
        print(f"DEBUG: Applied group filter, count: {institutions.count()}")
    if selected_states:
        institutions = institutions.filter(state__in=selected_states)
        print(f"DEBUG: Applied state filter, count: {institutions.count()}")
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        institutions = institutions.filter(is_active__in=active_status)
        print(f"DEBUG: Applied status filter, count: {institutions.count()}")

    # Get total count for pagination info
    total_count = institutions.count()
    print(f"DEBUG: Total institutions after filtering: {total_count}")
    
    # Apply pagination
    institutions = institutions[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for institution in institutions:
        data.append({
            'id': institution.pk,
            'name': institution.name,
            'address': institution.address,
            'district': institution.district,
            'state': institution.state,
            'total_strength': institution.total_strength,
            'is_active': institution.is_active,
            'group': institution.group.name if institution.group else None,
            'edit_url': reverse('institution_edit', args=[institution.pk]),
            'delete_url': reverse('institution_delete', args=[institution.pk]),
            'onboarding_type': institution.onboarding_type,
        })
    
    print(f"DEBUG: Returning {len(data)} institutions")
    
    return JsonResponse({
        'institutions': data,
        'total_count': total_count,
        'all_loaded': all_loaded,
        'search_query': search_query,
    })

# Hospital Views
@login_required
def hospital_list(request):
    # Sync strength counts from Firebase before loading the list
    if not (request.user.has_all_permissions() or 'view_hospitals' in request.user.get_all_permissions()):
        return redirect('base')
    try:
        print("DEBUG: Syncing strength counts before loading hospital list")
        sync_strength_counts(request)
    except Exception as e:
        print(f"DEBUG: Error syncing strength counts: {str(e)}")
    
    hospitals = Hospital.objects.all().order_by('-created_at') if request.user.has_all_permissions() else request.user.assigned_hospitals.all().order_by('-created_at')

    # Get all unique groups for the filter dropdown
    all_groups = Group.objects.filter(is_active=True, type="hospital").values('id', 'name')
    
    # Get all unique states for the filter dropdown
    all_states = Hospital.objects.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='') if request.user.has_all_permissions() else request.user.assigned_hospitals.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='')
    
    # Filtering
    search_query = request.GET.get('query', '').strip()
    selected_groups = request.GET.getlist('group')
    selected_states = request.GET.getlist('state')
    selected_statuses = request.GET.getlist('status')
    
    if search_query:
        hospitals = hospitals.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(district__icontains=search_query) |
            Q(state__icontains=search_query)
        )
    
    if selected_groups:
        hospitals = hospitals.filter(group_id__in=selected_groups)
    
    if selected_states:
        hospitals = hospitals.filter(state__in=selected_states)
    
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        hospitals = hospitals.filter(is_active__in=active_status)
    
    paginator = Paginator(hospitals, 10)
    page = request.GET.get('page')
    hospitals = paginator.get_page(page)
    
    return render(request, 'assessments/onboarding/hospital_list.html', {
        'hospitals': hospitals,
        'all_groups': all_groups,
        'all_states': all_states,
        'selected_groups': selected_groups,
        'selected_states': selected_states,
        'selected_statuses': selected_statuses,
        'search_query': search_query
    })

@login_required
def hospital_create(request):
    if not (request.user.has_all_permissions() or 'create_hospital' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    if request.method == 'POST':
        form = HospitalForm(request.POST)
        if form.is_valid():
            hospital_name = form.cleaned_data['name'].strip()
            onboarding_type = str(form.cleaned_data.get('onboarding_type', '')).lower()
            user = None
            # Validate uniqueness by onboarding type and name (case-insensitive)
            if Hospital.objects.filter(name__iexact=hospital_name, onboarding_type=onboarding_type).exists():
                return JsonResponse({'error': 'Hospital already exists for this onboarding type.'}, status=400)

            hospital = form.save(commit=False)
            if request.POST.get('is_active') == 'on':
                hospital.is_active = True
            else:
                hospital.is_active = False
            hospital.save()
            messages.success(request, 'Hospital created successfully.')
            return HttpResponse('OK')
    else:
        form = HospitalForm()
    return render(request, 'assessments/onboarding/hospital_form.html', {'form': form, 'action': 'Create'})

@login_required
def hospital_edit(request, pk):
    if not (request.user.has_all_permissions() or 'edit_hospital' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    hospital = get_object_or_404(Hospital, pk=pk)
    hospital_old_name = hospital.name
    if request.method == 'POST':
        form = HospitalForm(request.POST, instance=hospital)
        if form.is_valid():
            hospital_name = form.cleaned_data['name'].strip()
            onboarding_type = str(form.cleaned_data.get('onboarding_type', '')).lower()
            # Enforce uniqueness within same onboarding type (exclude current)
            if Hospital.objects.filter(name__iexact=hospital_name, onboarding_type=onboarding_type).exclude(pk=pk).exists():
                return JsonResponse({'error': 'Hospital already exists for this onboarding type.'}, status=400)
            
            
            if request.POST.get('is_active') == 'on':
                hospital.is_active = True
            else:
                hospital.is_active = False
            hospital = form.save(commit=False)
            hospital.save()
            messages.success(request, 'Hospital updated successfully.')
            return HttpResponse('OK')
    else:
        form = HospitalForm(instance=hospital)
    return render(request, 'assessments/onboarding/hospital_form.html', {'form': form, 'action': 'Edit'})

@login_required
def hospital_delete(request, pk):
    if not (request.user.has_all_permissions() or 'delete_hospital' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    hospital = get_object_or_404(Hospital, pk=pk)
    if request.method == 'POST':
        hospital.delete()
        messages.success(request, 'Hospital deleted successfully.')
        return redirect('hospital_list')
    return redirect('hospital_list')

@login_required
def hospital_list_api(request):
    # Sync strength counts from Firebase before loading the data
    if not (request.user.has_all_permissions() or 'view_hospitals' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    try:
        print("DEBUG: Syncing strength counts before loading hospital API data")
        sync_strength_counts(request)
    except Exception as e:
        print(f"DEBUG: Error syncing strength counts: {str(e)}")
    
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()
    
    print(f"DEBUG: Hospital search query received: '{search_query}'")
    
    hospitals = Hospital.objects.all().order_by('-created_at') if request.user.has_all_permissions() else request.user.assigned_hospitals.all().order_by('-created_at')
    
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying hospital search filter for: '{search_query}'")
        hospitals = hospitals.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(district__icontains=search_query) |
            Q(state__icontains=search_query)
        )
    
    # Apply other filters
    selected_groups = request.GET.getlist('group')
    selected_states = request.GET.getlist('state')
    selected_statuses = request.GET.getlist('status')

    print(f"DEBUG: Group filters received: {selected_groups}")
    print(f"DEBUG: State filters received: {selected_states}")
    print(f"DEBUG: Status filters received: {selected_statuses}")

    if selected_groups:
        hospitals = hospitals.filter(group_id__in=selected_groups)
        print(f"DEBUG: Applied group filter, count: {hospitals.count()}")
    if selected_states:
        hospitals = hospitals.filter(state__in=selected_states)
        print(f"DEBUG: Applied state filter, count: {hospitals.count()}")
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        hospitals = hospitals.filter(is_active__in=active_status)
        print(f"DEBUG: Applied status filter, count: {hospitals.count()}")

    # Get total count for pagination info
    total_count = hospitals.count()
    print(f"DEBUG: Total hospitals after filtering: {total_count}")
    
    # Apply pagination
    hospitals = hospitals[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for hospital in hospitals:
        data.append({
            'id': hospital.pk,
            'name': hospital.name,
            'address': hospital.address,
            'district': hospital.district,
            'state': hospital.state,
            'nurse_strength': hospital.nurse_strength,
            'is_active': hospital.is_active,
            'group': hospital.group.name if hospital.group else None,
            'edit_url': reverse('hospital_edit', args=[hospital.pk]),
            'delete_url': reverse('hospital_delete', args=[hospital.pk]),
        })
    
    print(f"DEBUG: Returning {len(data)} hospitals")
    
    return JsonResponse({
        'hospitals': data,
        'total_count': total_count,
        'all_loaded': all_loaded,
        'search_query': search_query,
    })

@login_required
def group_list_api(request):
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()
    
    print(f"DEBUG: Group search query received: '{search_query}'")
    
    groups = Group.objects.all().order_by('-created_at')
    
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying group search filter for: '{search_query}'")
        groups = groups.filter(
            Q(name__icontains=search_query) |
            Q(group_head__full_name__icontains=search_query) |
            Q(group_head__email__icontains=search_query)
        )
        print(f"DEBUG: Filtered groups count: {groups.count()}")
    
    # Apply other filters
    selected_types = request.GET.getlist('type')
    selected_statuses = request.GET.getlist('status')

    if selected_types:
        groups = groups.filter(type__in=selected_types)
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        groups = groups.filter(is_active__in=active_status)

    # Get total count for pagination info
    total_count = groups.count()
    
    # Apply pagination
    groups = groups[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for group in groups:
        data.append({
            'id': group.pk,
            'name': group.name,
            'type': group.get_type_display(),
            'unit_count': group.institution_set.count() if group.type == 'institution' else group.hospital_set.count(),
            'group_head': group.group_head.get_full_name() if group.group_head else '-',
            'group_head_email': group.group_head.email if group.group_head else '-',
            'group_head_phone': group.group_head.phone_number if group.group_head else '-',
            'is_active': group.is_active,
            'edit_url': reverse('group_edit', args=[group.pk]),
            'delete_url': reverse('group_delete', args=[group.pk]),
        })
    
    print(f"DEBUG: Returning {len(data)} groups")
    
    return JsonResponse({
        'groups': data,
        'all_loaded': all_loaded,
        'total_count': total_count,
        'offset': offset,
        'limit': limit,
        'search_query': search_query
    })

# Learner Views
@login_required
def learner_list(request):
    if not (request.user.has_all_permissions() or 'view_learners' in request.user.get_all_permissions()):
        return redirect('base')
    return render(request, 'assessments/onboarding/learner_list.html', {
        'institutions': Institution.objects.all(),
        'hospitals': Hospital.objects.all(),
        'selected_institutions': [],
        'selected_hospitals': [],
        'selected_learner_types': [],
    })

@login_required
def learner_create(request):
    if not (request.user.has_all_permissions() or 'add_learner' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    if request.method == 'POST':
        form = LearnerForm(request.POST)
        if form.is_valid():
            # Get form data
            learner_email = form.cleaned_data['learner_email']
            onboarding_type = form.cleaned_data.get('onboarding_type', '').lower()
            
            # Check for duplicate email with same onboarding type
            existing_learner = Learner.objects.filter(
                learner_user__email=learner_email,
                onboarding_type=onboarding_type
            ).first()
            
            if existing_learner:
                return JsonResponse({
                    'error': f'A learner with email {learner_email} and onboarding type {onboarding_type.upper()} already exists.',
                    'field': 'learner_email'
                }, status=400)
            
            # Enforce B2C requirements by learner type
            if str(onboarding_type).lower() == 'b2c':
                learner_type = form.cleaned_data.get('learner_type', '').lower()
                skillathon = form.cleaned_data.get('skillathon_event')
                college = form.cleaned_data.get('college')
                hospital = form.cleaned_data.get('hospital')
                if learner_type == 'student':
                    if not skillathon or not college:
                        return JsonResponse({
                            'error': 'For B2C Students, Skillathon Event and College are required.'
                        }, status=400)
                elif learner_type == 'nurse':
                    if not skillathon or not hospital:
                        return JsonResponse({
                            'error': 'For B2C Working Nurses, Skillathon Event and Hospital are required.'
                        }, status=400)
            learner_name = form.cleaned_data['learner_name']
            learner_phone = form.cleaned_data['learner_phone']

            if learner_name == '' or learner_email == '' or learner_phone == '':
                pass
            else:
                user, created = User.objects.get_or_create(
                    email=learner_email,
                    defaults={
                        'is_active': True,
                    }
                )
                
                if created:
                    user.full_name = learner_name
                    user.phone_number = learner_phone
                    default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                    user.set_password(default_password)
                    user.save()
                    # reset_link = request.build_absolute_uri(reverse('login'))
                    # subject = 'Your Learner Account Created'
                    # body = f"""Dear {learner_name},\n\nYour learner account has been created.\n\nUsername: {learner_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                    # send_email(subject, body, [learner_email])
                else:
                    # Update name/phone if changed
                    updated = False
                    if user.full_name != learner_name:
                        user.full_name = learner_name
                        updated = True
                    if user.phone_number != learner_phone:
                        user.phone_number = learner_phone
                        updated = True
                    if updated:
                        user.save()
            
            learner = form.save(commit=False)
            if user is not None:
                learner.learner_user = user
            learner.save()

            # Create scheduler object for exam assignments if skillathon is assigned
            if learner.skillathon_event and learner.learner_user:
                import json
                scheduler_data = {
                    "learner_ids": [learner.learner_user.id],
                    "skillathon_name": learner.skillathon_event.name
                }
                SchedularObject.objects.create(
                    data=json.dumps(scheduler_data),
                    is_completed=False
                )

            messages.success(request, 'Learner created successfully.')
            return HttpResponse('OK')
    else:
        form = LearnerForm()
    return render(request, 'assessments/onboarding/learner_form.html', {'form': form, 'action': 'Create'})

@login_required
def learner_edit(request, pk):
    if not (request.user.has_all_permissions() or 'edit_learner' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    learner = get_object_or_404(Learner, pk=pk)
    if request.method == 'POST':
        form = LearnerForm(request.POST, instance=learner)
        if form.is_valid():
            # Get form data
            learner_email = form.cleaned_data['learner_email']
            onboarding_type = form.cleaned_data.get('onboarding_type', '').lower()
            
            # Check for duplicate email with same onboarding type (excluding current learner)
            existing_learner = Learner.objects.filter(
                learner_user__email=learner_email,
                onboarding_type=onboarding_type
            ).exclude(pk=pk).first()
            
            if existing_learner:
                return JsonResponse({
                    'error': f'A learner with email {learner_email} and onboarding type {onboarding_type.upper()} already exists.',
                    'field': 'learner_email'
                }, status=400)
            
            # Enforce B2C requirements on update by learner type
            print(form.cleaned_data)
            if str(onboarding_type).lower() == 'b2c':
                learner_type = form.cleaned_data.get('learner_type', '').lower()
                skillathon = form.cleaned_data.get('skillathon_event')
                college = form.cleaned_data.get('college')
                hospital = form.cleaned_data.get('hospital')
                if learner_type == 'student':
                    if not skillathon or not college:
                        return JsonResponse({
                            'error': 'For B2C Students, Skillathon Event and College are required.'
                        }, status=400)
                elif learner_type == 'nurse':
                    if not skillathon or not hospital:
                        return JsonResponse({
                            'error': 'For B2C Working Nurses, Skillathon Event and Hospital are required.'
                        }, status=400)
            learner_name = form.cleaned_data['learner_name']
            learner_phone = form.cleaned_data['learner_phone']
            current_user = learner.learner_user
            
            if not current_user or current_user.email != learner_email:
                user, created = User.objects.get_or_create(
                    email=learner_email,
                    defaults={
                        'is_active': True,
                    }
                )
                if created:
                    user.full_name = learner_name
                    user.phone_number = learner_phone
                    default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                    user.set_password(default_password)
                    user.save()
                    # reset_link = request.build_absolute_uri(reverse('login'))
                    # subject = 'Your Learner Account Created'
                    # body = f"""Dear {learner_name},\n\nYour learner account has been created.\n\nUsername: {learner_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                    # send_email(subject, body, [learner_email])
                else:
                    updated = False
                    if user.full_name != learner_name:
                        user.full_name = learner_name
                        updated = True
                    if user.phone_number != learner_phone:
                        user.phone_number = learner_phone
                        updated = True
                    if updated:
                        user.save()
            else:
                updated = False
                if current_user.full_name != learner_name:
                    current_user.full_name = learner_name
                    updated = True
                if current_user.phone_number != learner_phone:
                    current_user.phone_number = learner_phone
                    updated = True
                if updated:
                    current_user.save()
            
            learner = form.save(commit=False)
            learner.save()

            # Create scheduler object for exam assignments if skillathon is assigned
            if learner.skillathon_event and learner.learner_user:
                import json
                scheduler_data = {
                    "learner_ids": [learner.learner_user.id],
                    "skillathon_name": learner.skillathon_event.name
                }
                SchedularObject.objects.create(
                    data=json.dumps(scheduler_data),
                    is_completed=False
                )

            messages.success(request, 'Learner updated successfully.')
            return HttpResponse('OK')
    else:
        form = LearnerForm(instance=learner, initial={
            'learner_name': learner.learner_user.full_name,
            'learner_email': learner.learner_user.email,
            'learner_phone': learner.learner_user.phone_number,
        })
    return render(request, 'assessments/onboarding/learner_form.html', {'form': form, 'action': 'Edit'})

@login_required
def learner_bulk_upload(request):
    if not (request.user.has_all_permissions() or 'bulk_upload_learners' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded'
            })
            
        if not file.name.endswith('.xlsx'):
            return JsonResponse({
                'success': False,
                'error': 'Please upload a valid Excel (.xlsx) file'
            })

        try:
            # Save file temporarily
            file_name = f"{uuid.uuid4()}_{file.name}".replace(" ", "_")
            file_path = os.path.join('media', 'uploaded_excels', file_name)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            # ===== BASIC FILE STRUCTURE VALIDATION ONLY =====
            print(f"[DEBUG] Checking Excel file structure...")
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Define required headers
            required_headers = [
                'Onboarding Type', 'Learner Type', 'Learner Name', 'Learner Email',
                'Learner Phone', 'College', 'Hospital', 'Skillathon Event',
                'Learner Gender'
            ]

            # Read headers
            headers = [cell.value for cell in ws[1]]

            # Only validate that required columns exist
            missing_headers = []
            for required_col in required_headers:
                if required_col not in headers:
                    missing_headers.append(required_col)

            if missing_headers:
                os.remove(file_path)  # Clean up file
                return JsonResponse({
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_headers)}'
                })

            # Get total row count for progress tracking
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty_rows = [
                row for row in all_rows
                if any(cell is not None and str(cell).strip() for cell in row)
            ]
            total_rows = len(non_empty_rows)

            # ===== FILE ACCEPTED - START BACKGROUND PROCESSING =====
            session_key = str(uuid.uuid4())

            # Initialize session data for progress tracking
            from datetime import datetime
            upload_session = {
                'session_key': session_key,
                'filename': file.name,
                'status': 'validating',
                'uploaded_by': request.user.email,
                'uploaded_at': datetime.now().isoformat(),
                'total_rows': total_rows,
                'processed_rows': 0,
                'success_count': 0,
                'error_count': 0,
                'created_count': 0,
                'row_results': []  # Will store {row_number, status, name, email, errors[]}
            }
            cache.set(f"upload_session:{session_key}", upload_session, timeout=7200)  # 2 hours

            # Add to active sessions list
            active_sessions = cache.get('active_upload_sessions', [])
            active_sessions.append(session_key)
            cache.set('active_upload_sessions', active_sessions, timeout=7200)

            # Start background processing
            print(f"[DEBUG] File accepted. Starting background validation for session: {session_key}")
            background_thread = threading.Thread(
                target=process_bulk_upload_with_progress,
                args=(file_path, session_key),
                daemon=True
            )
            background_thread.start()

            return JsonResponse({
                'success': True,
                'session_key': session_key,
                'message': 'Thank you for uploading! We will start with the validation.',
                'filename': file.name,
                'total_rows': total_rows
            })

        except Exception as e:
            print(e)
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'error': f'Error processing file: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

def process_bulk_upload_with_progress(file_path, session_key):
    """
    Validate and process bulk upload in background, tracking row-wise results
    """
    print(f"[DEBUG] Starting background validation for session: {session_key}")

    try:
        # Load Excel file
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Read headers
        headers = [cell.value for cell in ws[1]]

        # Get all rows and filter out empty ones
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        non_empty_rows = [
            (idx, row) for idx, row in enumerate(all_rows, start=2)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        print(f"[DEBUG] Validating {len(non_empty_rows)} rows...")

        # ===== PHASE 1: VALIDATION =====
        # Get all reference data from database
        all_skillathons = set(SkillathonEvent.objects.values_list('name', flat=True))
        all_colleges = set(Institution.objects.values_list('name', flat=True))
        all_hospitals = set(Hospital.objects.values_list('name', flat=True))
        existing_emails = set(EbekUser.objects.values_list('email', flat=True))

        # Helper function to safely get string value
        def safe_str(value):
            if value is None:
                return ''
            return str(value).strip()

        # Helper function to validate email format
        def is_valid_email(email):
            import re
            pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            return re.match(pattern, email) is not None

        # Validate each row and store results
        seen_emails = set()
        row_results = []

        for row_idx, row in non_empty_rows:
            row_data = dict(zip(headers, row))
            row_errors = []

            # Extract basic info
            learner_name = safe_str(row_data.get('Learner Name'))
            email = safe_str(row_data.get('Learner Email'))
            learner_type = safe_str(row_data.get('Learner Type')).lower()

            # Validate Onboarding Type
            onboarding_type = safe_str(row_data.get('Onboarding Type')).upper()
            if onboarding_type not in ['B2B', 'B2C']:
                row_errors.append(f"Onboarding Type must be 'B2B' or 'B2C', found: '{row_data.get('Onboarding Type', '')}'")

            # Validate Learner Name
            if not learner_name:
                row_errors.append("Learner Name cannot be empty")

            # Validate Email
            if not email:
                row_errors.append("Learner Email cannot be empty")
            else:
                if not is_valid_email(email):
                    row_errors.append(f"Invalid email format: '{email}'")
                if email in seen_emails:
                    row_errors.append(f"Duplicate email address in file: '{email}'")
                else:
                    seen_emails.add(email)
                if email in existing_emails:
                    row_errors.append(f"Email address already exists in system: '{email}'")

            # Validate Phone Number
            phone = safe_str(row_data.get('Learner Phone'))
            if not phone:
                row_errors.append("Learner Phone cannot be empty")

            # Validate Learner Gender
            gender = safe_str(row_data.get('Learner Gender'))
            if gender and gender.lower() not in ['male', 'female', 'other']:
                row_errors.append(f"Learner Gender must be 'Male', 'Female', or 'Other', found: '{gender}'")

            # Validate Learner Type
            if learner_type not in ['student', 'nurse']:
                row_errors.append(f"Learner Type must be 'Student' or 'Nurse', found: '{row_data.get('Learner Type', '')}'")

            # Validate Skillathon Event
            skillathon = safe_str(row_data.get('Skillathon Event'))
            if skillathon and skillathon not in all_skillathons:
                row_errors.append(f"Skillathon Event '{skillathon}' does not exist. Please create it first before uploading.")

            # Validate College for students
            if learner_type == 'student':
                college = safe_str(row_data.get('College'))
                if not college:
                    row_errors.append("College cannot be empty for students")
                elif college not in all_colleges:
                    row_errors.append(f"College '{college}' does not exist in system. Please create it first before uploading.")

            # Validate Hospital for nurses
            if learner_type == 'nurse':
                hospital = safe_str(row_data.get('Hospital'))
                if not hospital:
                    row_errors.append("Hospital cannot be empty for nurses")
                elif hospital not in all_hospitals:
                    row_errors.append(f"Hospital '{hospital}' does not exist in system. Please create it first before uploading.")

            # Store row result
            row_result = {
                'row_number': row_idx,
                'name': learner_name,
                'email': email,
                'learner_type': learner_type,
                'status': 'fail' if row_errors else 'pass',
                'errors': row_errors
            }
            row_results.append(row_result)

        # Update session with validation results
        session_data = cache.get(f"upload_session:{session_key}", {})
        if session_data:
            session_data['status'] = 'validated'
            session_data['row_results'] = row_results
            session_data['processed_rows'] = len(row_results)
            session_data['success_count'] = sum(1 for r in row_results if r['status'] == 'pass')
            session_data['error_count'] = sum(1 for r in row_results if r['status'] == 'fail')
            cache.set(f"upload_session:{session_key}", session_data, timeout=7200)

        print(f"[DEBUG] Validation complete: {session_data['success_count']} passed, {session_data['error_count']} failed")

        # ===== CHECK: ALL-OR-NOTHING VALIDATION =====
        # If even 1 row fails, reject entire upload
        if session_data['error_count'] > 0:
            session_data = cache.get(f"upload_session:{session_key}", {})
            if session_data:
                session_data['status'] = 'completed'
                session_data['message'] = f'Upload rejected: {session_data["error_count"]} row(s) failed validation. Please fix all errors and re-upload. No learners were created.'
                cache.set(f"upload_session:{session_key}", session_data, timeout=7200)
            logger.info(f"[DEBUG] Upload rejected due to validation errors: {session_data['error_count']} row(s) failed")
            return

        # ===== PHASE 2: CREATE LEARNERS FOR PASSED ROWS =====
        if session_data['success_count'] > 0:
            session_data['status'] = 'creating'
            cache.set(f"upload_session:{session_key}", session_data, timeout=7200)

            # Define header mapping for form data
            header_map = {
                'Onboarding Type': 'onboarding_type',
                'Learner Type': 'learner_type',
                'Learner Name': 'learner_name',
                'Learner Email': 'learner_email',
                'Learner Phone': 'learner_phone',
                'College': 'college',
                'Course': 'course',
                'Stream': 'stream',
                'Year of Study': 'year_of_study',
                'Hospital': 'hospital',
                'Designation': 'designation',
                'Years of Experience': 'years_of_experience',
                'Educational Qualification': 'educational_qualification',
                'Speciality': 'speciality',
                'State': 'state',
                'District': 'district',
                'Pincode': 'pincode',
                'Address': 'address',
                'Date of Birth': 'date_of_birth',
                'Certifications': 'certifications',
                'Learner Gender': 'learner_gender',
                'Skillathon Event': 'skillathon_event',
                'Educational Institution': 'educational_institution',
            }

            User = get_user_model()
            created_count = 0
            skillathon_learner_ids = []
            skillathon_name = ""

            # Process only rows that passed validation
            print(f"[DEBUG] Starting to create learners for {len(non_empty_rows)} total rows")
            print(f"[DEBUG] Row results count: {len(row_results)}")

            for row_idx, row in non_empty_rows:
                # Find the result for this row
                row_result = next((r for r in row_results if r['row_number'] == row_idx), None)
                if not row_result:
                    print(f"[DEBUG] Row {row_idx}: No validation result found")
                    continue
                if row_result['status'] != 'pass':
                    print(f"[DEBUG] Row {row_idx}: Validation failed, skipping")
                    continue

                print(f"[DEBUG] Row {row_idx}: Starting creation process")
                try:
                    row_data = dict(zip(headers, row))

                    # Prepare form data
                    form_data = {}
                    for excel_col, form_field in header_map.items():
                        value = row_data.get(excel_col)
                        if value is not None:
                            # Convert to string first (handles both strings and numbers)
                            value_str = str(value).strip()
                            if value_str:
                                # Handle ForeignKeys by name
                                if form_field == 'college' and value_str:
                                    try:
                                        form_data['college'] = Institution.objects.get(name=value_str).id
                                    except Institution.DoesNotExist:
                                        form_data['college'] = None
                                elif form_field == 'hospital' and value_str:
                                    try:
                                        form_data['hospital'] = Hospital.objects.get(name=value_str).id
                                    except Hospital.DoesNotExist:
                                        form_data['hospital'] = None
                                else:
                                    form_data[form_field] = value_str

                                # Convert to lowercase
                                if form_field == 'onboarding_type':
                                    form_data['onboarding_type'] = value_str.lower()
                                if form_field == 'learner_type':
                                    form_data['learner_type'] = value_str.lower()
                                if form_field == 'learner_gender':
                                    form_data['learner_gender'] = value_str.lower()
                                if form_field == 'skillathon_event':
                                    try:
                                        form_data['skillathon_event'] = SkillathonEvent.objects.get(name=value_str).id
                                        skillathon_name = value_str
                                    except SkillathonEvent.DoesNotExist:
                                        form_data['skillathon_event'] = None

                    # Add required user fields (convert to string to handle both text and numbers)
                    learner_name = row_data.get('Learner Name', '')
                    form_data['learner_name'] = str(learner_name).strip() if learner_name else ''

                    learner_email = row_data.get('Learner Email', '')
                    form_data['learner_email'] = str(learner_email).strip() if learner_email else ''

                    learner_phone = row_data.get('Learner Phone', '')
                    form_data['learner_phone'] = str(learner_phone).strip()[:10] if learner_phone else ''

                    form = LearnerForm(form_data)
                    if form.is_valid():
                        # Create new learner
                        email = form.cleaned_data['learner_email']
                        full_name = form.cleaned_data['learner_name']
                        phone = form.cleaned_data['learner_phone']
                        user, user_created = User.objects.get_or_create(
                            email=email,
                            defaults={'full_name': full_name, 'phone_number': phone, 'is_active': True}
                        )
                        if user_created:
                            default_password = 'Success@123$'
                            user.set_password(default_password)
                            user.save()

                        learner = form.save(commit=False)
                        learner.learner_user = user
                        learner.save()
                        created_count += 1

                        # Track for skillathon assignments
                        if learner.skillathon_event:
                            skillathon_learner_ids.append(user.id)

                        print(f"[DEBUG] Created learner: {full_name} ({created_count}/{session_data['success_count']})")

                        # Update progress in cache after each learner creation
                        progress_data = cache.get(f"upload_session:{session_key}", {})
                        if progress_data:
                            progress_data['created_count'] = created_count
                            progress_data['message'] = f'Creating learner {created_count} of {session_data["success_count"]}: {full_name}'
                            cache.set(f"upload_session:{session_key}", progress_data, timeout=7200)
                    else:
                        print(f"[DEBUG] Row {row_idx}: Form validation failed")
                        print(f"[DEBUG] Form errors: {form.errors}")
                        logger.error(f"Form validation failed for row {row_idx}: {form.errors}")

                except Exception as e:
                    logger.error(f"Error creating learner for row {row_idx}: {str(e)}")
                    logger.error(traceback.format_exc())

            # Create scheduler object for skillathon assignments
            if skillathon_learner_ids and skillathon_name:
                logger.info(f"[DEBUG] Creating scheduler for {len(skillathon_learner_ids)} learners with skillathon: {skillathon_name}")
                SchedularObject.objects.create(
                    data=json.dumps({
                        "learner_ids": skillathon_learner_ids,
                        "skillathon_name": skillathon_name
                    }),
                    is_completed=False
                )

            # Update final status
            session_data = cache.get(f"upload_session:{session_key}", {})
            if session_data:
                session_data['status'] = 'completed'
                session_data['message'] = f'Validation complete. {created_count} learner(s) created successfully. {session_data["error_count"]} row(s) failed validation.'
                cache.set(f"upload_session:{session_key}", session_data, timeout=7200)

            logger.info(f"[DEBUG] Bulk upload completed: {created_count} created, {session_data['error_count']} failed")
        else:
            # No valid rows to create
            session_data = cache.get(f"upload_session:{session_key}", {})
            if session_data:
                session_data['status'] = 'completed'
                session_data['message'] = f'Validation complete. All {session_data["error_count"]} row(s) failed validation. No learners created.'
                cache.set(f"upload_session:{session_key}", session_data, timeout=7200)

    except Exception as e:
        # Handle any unexpected errors
        session_data = cache.get(f"upload_session:{session_key}", {})
        if session_data:
            session_data['status'] = 'error'
            session_data['message'] = f'Error processing file: {str(e)}'
            cache.set(f"upload_session:{session_key}", session_data, timeout=7200)
        logger.error(f"Bulk upload error: {e}")
        logger.error(traceback.format_exc())

@login_required
def get_active_upload_sessions(request):
    """
    API endpoint to get all active upload sessions for displaying progress bars.
    Returns basic info for all ongoing uploads.
    """
    active_session_keys = cache.get('active_upload_sessions', [])
    sessions = []

    for session_key in active_session_keys:
        session_data = cache.get(f"upload_session:{session_key}")
        if session_data:
            # Include all statuses (validating, validated, creating, completed, error)
            status = session_data.get('status', 'validating')
            total_rows = session_data.get('total_rows', 0)
            processed_rows = session_data.get('processed_rows', 0)

            # Calculate progress percentage
            if status in ['validating', 'validated', 'creating']:
                # Show progress based on processed rows
                progress_percentage = int((processed_rows / total_rows * 100)) if total_rows > 0 else 0
            elif status == 'completed':
                progress_percentage = 100
            else:
                progress_percentage = 0

            sessions.append({
                'session_key': session_key,
                'filename': session_data.get('filename', ''),
                'status': status,
                'uploaded_by': session_data.get('uploaded_by', ''),
                'uploaded_at': session_data.get('uploaded_at', ''),
                'total_rows': total_rows,
                'processed_rows': processed_rows,
                'success_count': session_data.get('success_count', 0),
                'error_count': session_data.get('error_count', 0),
                'progress_percentage': progress_percentage,
                'message': session_data.get('message', '')
            })

    return JsonResponse({
        'success': True,
        'sessions': sessions
    })

@login_required
def get_upload_session_details(request, session_key):
    """
    API endpoint to get detailed info about a specific upload session.
    Returns row-wise validation results for display in modal.
    """
    session_data = cache.get(f"upload_session:{session_key}")

    if not session_data:
        return JsonResponse({
            'success': False,
            'error': 'No upload session found for this key'
        })

    return JsonResponse({
        'success': True,
        'session_key': session_key,
        'filename': session_data.get('filename', ''),
        'status': session_data.get('status', 'validating'),
        'uploaded_by': session_data.get('uploaded_by', ''),
        'uploaded_at': session_data.get('uploaded_at', ''),
        'total_rows': session_data.get('total_rows', 0),
        'processed_rows': session_data.get('processed_rows', 0),
        'success_count': session_data.get('success_count', 0),
        'error_count': session_data.get('error_count', 0),
        'created_count': session_data.get('created_count', 0),
        'row_results': session_data.get('row_results', []),
        'message': session_data.get('message', '')
    })

def cascade_delete_learner_data(learner):
    """
    Cascading delete for learner data in Firebase.
    Removes learner from Batches, and deletes all associated ExamAssignments
    and their references from BatchAssignment or ProcedureAssignment.
    """
    try:
        if not learner.learner_user:
            print(f"Warning: Learner {learner.pk} has no associated user")
            return

        user_id = learner.learner_user.id
        user_path = f'/Users/{user_id}'

        print(f"Starting cascading delete for learner user: {user_path}, onboarding_type: {learner.onboarding_type}")

        # Step 1: Find all ExamAssignments for this user
        exam_assignments_query = db.collection('ExamAssignment').where('user', '==', user_path).stream()
        exam_assignment_ids = []
        exam_assignment_paths = []

        for exam_doc in exam_assignments_query:
            exam_assignment_ids.append(exam_doc.id)
            exam_assignment_paths.append(f'/ExamAssignment/{exam_doc.id}')

        print(f"Found {len(exam_assignment_ids)} ExamAssignment(s) for user {user_path}")

        if learner.onboarding_type.lower() == 'b2b':
            # B2B: Remove from Batches, BatchAssignment, then delete ExamAssignments

            # Step 2a: Remove user from all Batches and collect batch paths
            batches_query = db.collection('Batches').where('learners', 'array_contains', user_path).stream()
            batch_count = 0
            batch_paths = []

            for batch_doc in batches_query:
                batch_id = batch_doc.id
                batch_path = f'/Batches/{batch_id}'
                batch_paths.append(batch_path)

                batch_ref = db.collection('Batches').document(batch_id)
                batch_ref.update({
                    'learners': firestore.ArrayRemove([user_path])
                })
                batch_count += 1
                print(f"Removed user from Batch: {batch_id} (path: {batch_path})")

            print(f"Removed user from {batch_count} Batch(es)")
            print(f"Batch paths to check: {batch_paths}")

            # Step 2b: Remove ExamAssignment references from BatchAssignment for relevant batches only
            if exam_assignment_paths and batch_paths:
                print(f"Looking for BatchAssignments with batch in: {batch_paths}")

                # Query BatchAssignments only for the batches the user was in
                for batch_path in batch_paths:
                    batch_assignments_query = db.collection('BatchAssignment').where('batch', '==', batch_path).stream()

                    for ba_doc in batch_assignments_query:
                        ba_data = ba_doc.to_dict()
                        ba_id = ba_doc.id
                        ba_ref = db.collection('BatchAssignment').document(ba_id)

                        print(f"Found BatchAssignment: {ba_id} for batch: {batch_path}")
                        print(f"  Current examassignment array: {[ref.path for ref in ba_data.get('examassignment', [])]}")

                        # Check each exam assignment reference
                        removed_count = 0
                        for exam_path in exam_assignment_paths:
                            if exam_path in ba_data.get('examassignment', []):
                                ba_ref.update({
                                    'examassignment': firestore.ArrayRemove([exam_path])
                                })
                                removed_count += 1
                                print(f"  Removed {exam_path} from BatchAssignment: {ba_id}")

                        if removed_count == 0:
                            print(f"  No matching exam assignments found in this BatchAssignment")
            elif exam_assignment_paths and not batch_paths:
                print(f"Warning: User has {len(exam_assignment_paths)} exam assignments but was not found in any batches")

        else:
            # B2C: Remove ExamAssignment references from ProcedureAssignment

            if exam_assignment_paths:
                procedure_assignments_query = db.collection('ProcedureAssignment').stream()
                for pa_doc in procedure_assignments_query:
                    pa_data = pa_doc.to_dict()
                    pa_ref = db.collection('ProcedureAssignment').document(pa_doc.id)

                    # Check each exam assignment reference
                    for exam_path in exam_assignment_paths:
                        if exam_path in pa_data.get('examAssignmentArray', []):
                            pa_ref.update({
                                'examAssignmentArray': firestore.ArrayRemove([exam_path])
                            })
                            print(f"Removed {exam_path} from ProcedureAssignment: {pa_doc.id}")

        # Step 3: Delete all ExamAssignments
        for exam_id in exam_assignment_ids:
            db.collection('ExamAssignment').document(exam_id).delete()
            print(f"Deleted ExamAssignment: {exam_id}")

        print(f"Cascading delete completed for learner user: {user_path}")

    except Exception as e:
        print(f"Error during cascading delete for learner {learner.pk}: {str(e)}")
        import traceback
        traceback.print_exc()

@login_required
def learner_delete(request, pk):
    learner = get_object_or_404(Learner, pk=pk)
    if request.method == 'POST':
        # Perform cascading delete before deleting the learner
        cascade_delete_learner_data(learner)
        learner.delete()
        messages.success(request, 'Learner deleted successfully.')
        return redirect('learner_list')
    return redirect('learner_list')

@login_required
def learners_bulk_delete(request):
    if not (request.user.has_all_permissions() or 'delete_learner' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    try:
        data = json.loads(request.body or '{}')
        ids = data.get('ids', [])
        if not ids:
            return JsonResponse({'error': 'No learners selected'}, status=400)
        qs = Learner.objects.filter(pk__in=ids)
        count = qs.count()
        for learner in qs:
            # Perform cascading delete before deleting the learner
            cascade_delete_learner_data(learner)
            learner.delete()
        return JsonResponse({'success': True, 'deleted': count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Assessor Views
@login_required
def assessor_list(request):
    if not (request.user.has_all_permissions() or 'view_assessors' in request.user.get_all_permissions()):
        return redirect('base')
    search_query = request.GET.get('search', '').strip()
    selected_specialities = request.GET.getlist('speciality')
    selected_statuses = request.GET.getlist('status')

    assigned_colleges = request.user.assigned_institutions.values_list('id', flat=True)
    assigned_hospitals = request.user.assigned_hospitals.values_list('id', flat=True)
    # Filtering logic (same as main view)
    if request.user.has_all_permissions():
        assessors = Assessor.objects.all().order_by('-created_at')
    else:
        assessors = Assessor.objects.filter(
            Q(institution_id__in=assigned_colleges) | Q(hospital_id__in=assigned_hospitals)
        ).order_by('-created_at')
    
    
    if search_query:
        assessors = assessors.filter(
            Q(assessor_user__full_name__icontains=search_query) |
            Q(assessor_user__email__icontains=search_query) |
            Q(assessor_user__phone_number__icontains=search_query) |
            Q(specialization__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    if selected_specialities:
        assessors = assessors.filter(specialization__in=selected_specialities)
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        assessors = assessors.filter(is_active__in=active_status)
    
    paginator = Paginator(assessors, 10)
    page = request.GET.get('page')
    assessors = paginator.get_page(page)
    
    all_specialities = Assessor.objects.values_list('specialization', flat=True).distinct()
    
    return render(request, 'assessments/onboarding/assessor_list.html', {
        'assessors': assessors,
        'all_specialities': all_specialities,
        'selected_specialities': selected_specialities,
        'selected_statuses': selected_statuses,
        'search_query': search_query,
    })

@login_required
def assessor_create(request):
    if not (request.user.has_all_permissions() or 'add_assessor' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    if request.method == 'POST':
        form = AssessorForm(request.POST)
        if form.is_valid():
            assessor_name = form.cleaned_data['assessor_name']
            assessor_email = form.cleaned_data['assessor_email']
            assessor_phone = form.cleaned_data['assessor_phone']
            
            # Check if assessor with same email already exists in Firebase
            existing_assessor_email = db.collection('Users').where('email', '==', assessor_email).limit(1).stream()
            if list(existing_assessor_email):
                return JsonResponse({'error': 'An assessor with this email already exists'}, status=400)
            
            # Check if assessor with same phone already exists in Firebase
            existing_assessor_phone = db.collection('Users').where('phone_number', '==', assessor_phone).limit(1).stream()
            if list(existing_assessor_phone):
                return JsonResponse({'error': 'An assessor with this phone number already exists'}, status=400)
            
            user, created = User.objects.get_or_create(
                email=assessor_email,
                defaults={
                    'is_active': True,
                }
            )
            
            if created:
                user.full_name = assessor_name
                user.phone_number = assessor_phone
                default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                user.set_password(default_password)
                user.save()
                reset_link = request.build_absolute_uri(reverse('login'))
                subject = 'Your Assessor Account Created'
                body = f"""Dear {assessor_name},\n\nYour assessor account has been created.\n\nUsername: {assessor_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                # send_email_thread = threading.Thread(target=send_email, args=(subject, body, [assessor_email]))
                # send_email_thread.start()
            else:
                # Update name/phone if changed
                updated = False
                if user.full_name != assessor_name:
                    user.full_name = assessor_name
                    updated = True
                if user.phone_number != assessor_phone:
                    user.phone_number = assessor_phone
                    updated = True
                if updated:
                    user.save()
            
            assessor = form.save(commit=False)
            assessor.assessor_user = user
            
            # Handle institution and hospital from custom select fields
            institution_id = request.POST.get('institution')
            hospital_id = request.POST.get('hospital')
            
            if institution_id:
                try:
                    assessor.institution = Institution.objects.get(id=institution_id)
                    
                except Institution.DoesNotExist:
                    pass
            
            if hospital_id:
                try:
                    assessor.hospital = Hospital.objects.get(id=hospital_id)
                except Hospital.DoesNotExist:
                    pass
            
            if request.POST.get('is_active') == 'on':
                assessor.is_active = True
            else:
                assessor.is_active = False
            
            assessor.save()
            assessor_user = assessor.assessor_user
            if institution_id:
                assessor_user.assigned_institutions.add(assessor.institution)
            if hospital_id:
                assessor_user.assigned_hospitals.add(assessor.hospital)
            assessor_user.save()
            
            messages.success(request, 'Assessor created successfully.')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return HttpResponse('OK')
            else:
                return redirect('assessor_list')
        else:
            # Form has validation errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return render(request, 'assessments/onboarding/assessor_form.html', {'form': form, 'action': 'Create'}, status=400)
    else:
        form = AssessorForm()
    
    return render(request, 'assessments/onboarding/assessor_form.html', {'form': form, 'action': 'Create'})

@login_required
def assessor_edit(request, pk):
    if not (request.user.has_all_permissions() or 'edit_assessor' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    User = get_user_model()
    assessor = get_object_or_404(Assessor, pk=pk)
    if request.method == 'POST':
        form = AssessorForm(request.POST, instance=assessor)
        if form.is_valid():
            assessor_name = form.cleaned_data['assessor_name']
            assessor_email = form.cleaned_data['assessor_email']
            assessor_phone = form.cleaned_data['assessor_phone']
            current_user = assessor.assessor_user

            # Check if email is being changed and if it already exists in Firebase
            if not current_user or current_user.email != assessor_email:
                existing_assessor_email = db.collection('Users').where('email', '==', assessor_email).limit(1).stream()
                if list(existing_assessor_email):
                    return JsonResponse({'error': 'An assessor with this email already exists'}, status=400)
            
            # Check if phone is being changed and if it already exists in Firebase
            if not current_user or current_user.phone_number != assessor_phone:
                existing_assessor_phone = db.collection('Users').where('phone_number', '==', assessor_phone).limit(1).stream()
                if list(existing_assessor_phone):
                    return JsonResponse({'error': 'An assessor with this phone number already exists'}, status=400)

            if assessor_name != "" and assessor_email != "" and assessor_phone != "":
                if not current_user or current_user.email != assessor_email:
                    user, created = User.objects.get_or_create(
                        email=assessor_email,
                        defaults={
                            'is_active': True,
                        }
                    )
                    if created:
                        user.full_name = assessor_name
                        user.phone_number = assessor_phone
                        default_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                        user.set_password(default_password)
                        user.save()
                        reset_link = request.build_absolute_uri(reverse('login'))
                        subject = 'Your Assessor Account Created'
                        body = f"""Dear {assessor_name},\n\nYour assessor account has been created.\n\nUsername: {assessor_email}\nPassword: {default_password}\n\nPlease log in to the platform using the link below and change your password after first login. Click the link to login: {reset_link}\n\nRegards,\nTeam"""
                        send_email_thread = threading.Thread(target=send_email, args=(subject, body, [assessor_email]))
                        send_email_thread.start()
                    else:
                        updated = False
                        if user.full_name != assessor_name:
                            user.full_name = assessor_name
                            updated = True
                        if user.phone_number != assessor_phone:
                            user.phone_number = assessor_phone
                            updated = True
                        if updated:
                            user.save()
                    assessor.assessor_user = user
                else:
                    updated = False
                    if current_user.full_name != assessor_name:
                        current_user.full_name = assessor_name
                        updated = True
                    if current_user.phone_number != assessor_phone:
                        current_user.phone_number = assessor_phone
                        updated = True
                    if updated:
                        current_user.save()
            
            # Handle institution and hospital from custom select fields
            institution_id = request.POST.get('institution')
            hospital_id = request.POST.get('hospital')
            
            if institution_id:
                try:
                    assessor.institution = Institution.objects.get(id=institution_id)
                except Institution.DoesNotExist:
                    pass
            else:
                assessor.institution = None
            
            if hospital_id:
                try:
                    assessor.hospital = Hospital.objects.get(id=hospital_id)
                except Hospital.DoesNotExist:
                    pass
            else:
                assessor.hospital = None
            
            if request.POST.get('is_active') == 'on':
                assessor.is_active = True
            else:
                assessor.is_active = False
            assessor = form.save(commit=False)
            assessor_user = assessor.assessor_user
            if institution_id:
                assessor_user.assigned_institutions.add(assessor.institution)
            if hospital_id:
                assessor_user.assigned_hospitals.add(assessor.hospital)
            assessor.save()
            assessor_user.save()
            messages.success(request, 'Assessor updated successfully.')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return HttpResponse('OK')
            else:
                return redirect('assessor_list')
        else:
            # Form has validation errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return render(request, 'assessments/onboarding/assessor_form.html', {'form': form, 'action': 'Edit'}, status=400)
    else:
        form = AssessorForm(instance=assessor, initial={
            'assessor_name': assessor.assessor_user.full_name,
            'assessor_email': assessor.assessor_user.email,
            'assessor_phone': assessor.assessor_user.phone_number,
        })
    
    return render(request, 'assessments/onboarding/assessor_form.html', {'form': form, 'action': 'Edit'})

@login_required
def assessor_delete(request, pk):
    if not (request.user.has_all_permissions() or 'delete_assessor' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    assessor = get_object_or_404(Assessor, pk=pk)
    if request.method == 'POST':
        assessor.delete()
        messages.success(request, 'Assessor deleted successfully.')
        return redirect('assessor_list')
    return redirect('assessor_list')

@login_required
def assessor_list_api(request):

    if not (request.user.has_all_permissions() or 'view_assessors' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()
    
    print(f"DEBUG: Assessor search query received: '{search_query}'")
    
    assigned_colleges = request.user.assigned_institutions.values_list('id', flat=True)
    assigned_hospitals = request.user.assigned_hospitals.values_list('id', flat=True)
    # Filtering logic (same as main view)
    if request.user.has_all_permissions():
        assessors = Assessor.objects.all().order_by('-created_at')
    else:
        assessors = Assessor.objects.filter(
            Q(institution_id__in=assigned_colleges) | Q(hospital_id__in=assigned_hospitals)
        ).order_by('-created_at')
        
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying assessor search filter for: '{search_query}'")
        assessors = assessors.filter(
            Q(assessor_user__full_name__icontains=search_query) |
            Q(assessor_user__email__icontains=search_query) |
            Q(assessor_user__phone_number__icontains=search_query) |
            Q(specialization__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    # Apply other filters
    selected_specialities = request.GET.getlist('speciality')
    selected_statuses = request.GET.getlist('status')

    print(f"DEBUG: Speciality filters received: {selected_specialities}")
    print(f"DEBUG: Status filters received: {selected_statuses}")

    if selected_specialities:
        if 'None' in selected_specialities:
            # Remove None to avoid issues in __in lookup
            cleaned_specialities = [s for s in selected_specialities if s is not None]
            assessors = assessors.filter(
                Q(specialization__isnull=True) | Q(specialization__in=cleaned_specialities)
            )
        else:
            assessors = assessors.filter(specialization__in=selected_specialities)
        print(f"DEBUG: Applied speciality filter, count: {assessors.count()}")
    if selected_statuses:
        active_status = [status.lower() == 'active' for status in selected_statuses]
        assessors = assessors.filter(is_active__in=active_status)
        print(f"DEBUG: Applied status filter, count: {assessors.count()}")

    # Get total count for pagination info
    total_count = assessors.count()
    print(f"DEBUG: Total assessors after filtering: {total_count}")
    
    # Apply pagination
    assessors = assessors[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for assessor in assessors:
        data.append({
            'id': assessor.pk,
            'full_name': assessor.assessor_user.get_full_name() if assessor.assessor_user else "Assessor",
            'email': assessor.assessor_user.email if assessor.assessor_user else "Assessor",
            'phone_number': assessor.assessor_user.phone_number if assessor.assessor_user else "Assessor",
            'specialization': assessor.specialization,
            'location': assessor.location,
            'is_active': assessor.is_active,
            'edit_url': reverse('assessor_edit', args=[assessor.pk]),
            'delete_url': reverse('assessor_delete', args=[assessor.pk]),
        })
    
    print(f"DEBUG: Returning {len(data)} assessors")
    
    return JsonResponse({
        'assessors': data,
        'total_count': total_count,
        'all_loaded': all_loaded,
        'search_query': search_query,
    })

@login_required
def get_institutions_hospitals(request):
    """API to get institutions and hospitals for assessor form"""
    unit_type = request.GET.get('unit_type')
    print(f"DEBUG: get_institutions_hospitals called with unit_type: {unit_type}")
    
    if unit_type == 'institution':
        institutions = Institution.objects.filter(is_active=True).values('id', 'name')
        print(f"DEBUG: Found {institutions.count()} active institutions")
        return JsonResponse({'institutions': list(institutions)})
    elif unit_type == 'hospital':
        hospitals = Hospital.objects.filter(is_active=True).values('id', 'name')
        print(f"DEBUG: Found {hospitals.count()} active hospitals")
        return JsonResponse({'hospitals': list(hospitals)})
    else:
        print(f"DEBUG: Invalid unit_type: {unit_type}")
        return JsonResponse({'institutions': [], 'hospitals': []})

# Add these new functions after the existing get_institutions_hospitals function

@login_required
def get_institutions_by_skillathon(request):
    """API to get institutions filtered by skillathon"""
    skillathon_id = request.GET.get('skillathon_id')
    onboarding_type = request.GET.get('onboarding_type', '').lower()
    
    print(f"DEBUG: get_institutions_by_skillathon called with skillathon_id: {skillathon_id}, onboarding_type: {onboarding_type}")
    
    institutions = Institution.objects.filter(is_active=True)
    
    # Filter by skillathon if provided
    if skillathon_id and skillathon_id != '':
        institutions = institutions.filter(skillathon_id=skillathon_id)
    
    # Filter by onboarding type if provided
    if onboarding_type:
        institutions = institutions.filter(onboarding_type=onboarding_type)
    
    institutions_data = list(institutions.values('id', 'name'))
    print(f"DEBUG: Found {len(institutions_data)} institutions")
    
    return JsonResponse({'institutions': institutions_data})

@login_required
def get_hospitals_by_skillathon(request):
    """API to get hospitals filtered by skillathon"""
    skillathon_id = request.GET.get('skillathon_id')
    onboarding_type = request.GET.get('onboarding_type', '').lower()
    
    print(f"DEBUG: get_hospitals_by_skillathon called with skillathon_id: {skillathon_id}, onboarding_type: {onboarding_type}")
    
    hospitals = Hospital.objects.filter(is_active=True)
    
    # Filter by skillathon if provided
    if skillathon_id and skillathon_id != '':
        hospitals = hospitals.filter(skillathon_id=skillathon_id)
    
    # Filter by onboarding type if provided
    if onboarding_type:
        hospitals = hospitals.filter(onboarding_type=onboarding_type)
    
    hospitals_data = list(hospitals.values('id', 'name'))
    print(f"DEBUG: Found {len(hospitals_data)} hospitals")
    
    return JsonResponse({'hospitals': hospitals_data})

# Skillathon Event Views
@login_required
def skillathon_list(request):
    if not (request.user.has_all_permissions() or 'view_skillathons' in request.user.get_all_permissions()):
        return redirect('base')
    
    assigned_skillathons = request.user.assigned_skillathons.values_list('id', flat=True)
    if request.user.has_all_permissions():
        events = SkillathonEvent.objects.all().order_by('-date')
    else:
        events = SkillathonEvent.objects.filter(id__in=assigned_skillathons).order_by('-date')
    
    events = SkillathonEvent.objects.all().order_by('-date')
    
    # Get all unique cities for the filter dropdown
    all_cities = SkillathonEvent.objects.values_list('city', flat=True).distinct().exclude(city__isnull=True).exclude(city='') if request.user.has_all_permissions() else request.user.assigned_skillathons.values_list('city', flat=True).distinct().exclude(city__isnull=True).exclude(city='')
    
    # Get all unique states for the filter dropdown
    all_states = SkillathonEvent.objects.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='') if request.user.has_all_permissions() else request.user.assigned_skillathons.values_list('state', flat=True).distinct().exclude(state__isnull=True).exclude(state='')
    
    # Filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    state = request.GET.get('state')
    
    if date_from:
        events = events.filter(date__gte=date_from)
    if date_to:
        events = events.filter(date__lte=date_to)
    if state:
        events = events.filter(state__icontains=state)
    
    paginator = Paginator(events, 10)
    page = request.GET.get('page')
    events = paginator.get_page(page)
    
    return render(request, 'assessments/onboarding/skillathon_list.html', {
        'events': events,
        'all_cities': all_cities,
        'all_states': all_states,
        'selected_cities': request.GET.getlist('city'),
        'selected_states': request.GET.getlist('state'),
        'date_from': date_from,
        'date_to': date_to,
        'state': state,
    })

@login_required
def skillathon_create(request):
    if not (request.user.has_all_permissions() or 'add_skillathon' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method == 'POST':
        form = SkillathonEventForm(request.POST)
        if form.is_valid():
            event = form.save()
            messages.success(request, 'Skillathon event created successfully.')
            return HttpResponse('OK')
    else:
        form = SkillathonEventForm()
    return render(request, 'assessments/onboarding/skillathon_form.html', {'form': form, 'action': 'Create'})

@login_required
def skillathon_edit(request, pk):
    if not (request.user.has_all_permissions() or 'edit_skillathon' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    event = get_object_or_404(SkillathonEvent, pk=pk)
    if request.method == 'POST':
        form = SkillathonEventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, 'Skillathon event updated successfully.')
            return HttpResponse('OK')
    else:
        form = SkillathonEventForm(instance=event)
    return render(request, 'assessments/onboarding/skillathon_form.html', {'form': form, 'action': 'Edit'}) 

@login_required
def skillathon_delete(request, pk):
    if not (request.user.has_all_permissions() or 'delete_skillathon' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    event = get_object_or_404(SkillathonEvent, pk=pk)
    if request.method == 'POST':
        event.delete()
        messages.success(request, 'Skillathon event deleted successfully.')
        return redirect('skillathon_list')
    return redirect('skillathon_list')

@login_required
def skillathon_list_api(request):
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()

    print(f"DEBUG: Skillathon search query received: '{search_query}'")
    
    assigned_skillathons = request.user.assigned_skillathons.values_list('id', flat=True)
    if request.user.has_all_permissions():
        events = SkillathonEvent.objects.all().order_by('-date')
    else:
        events = SkillathonEvent.objects.filter(id__in=assigned_skillathons).order_by('-date')
    
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying skillathon search filter for: '{search_query}'")
        events = events.filter(
            Q(name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(state__icontains=search_query)
        )
    
    # Apply date filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    print(f"DEBUG: Date filters - from: {date_from}, to: {date_to}")
    
    if date_from:
        events = events.filter(date__gte=date_from)
        print(f"DEBUG: Applied date_from filter, count: {events.count()}")
    if date_to:
        events = events.filter(date__lte=date_to)
        print(f"DEBUG: Applied date_to filter, count: {events.count()}")
    
    # Apply state filter
    state = request.GET.get('state')
    if state:
        events = events.filter(state__icontains=state)
        print(f"DEBUG: Applied state filter, count: {events.count()}")
    
    # Apply city filter
    selected_cities = request.GET.getlist('city')
    if selected_cities:
        events = events.filter(city__in=selected_cities)
        print(f"DEBUG: Applied city filter, count: {events.count()}")
    
    # Apply state filter (checkbox selection)
    selected_states = request.GET.getlist('state')
    if selected_states:
        events = events.filter(state__in=selected_states)
        print(f"DEBUG: Applied selected states filter, count: {events.count()}")
    
    # Get total count for pagination info
    total_count = events.count()
    print(f"DEBUG: Total skillathons after filtering: {total_count}")
    
    # Apply pagination
    events = events[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for event in events:
        data.append({
            'id': event.pk,
            'name': event.name,
            'date': event.date.strftime('%d %b %Y') if event.date else '-',
            'city': event.city,
            'state': event.state,
            'edit_url': reverse('skillathon_edit', args=[event.pk]),
            'delete_url': reverse('skillathon_delete', args=[event.pk]),
        })
    
    print(f"DEBUG: Returning {len(data)} skillathons")
    
    return JsonResponse({
        'events': data,
        'total_count': total_count,
        'all_loaded': all_loaded,
        'search_query': search_query,
    })

def reconnect_all_signals():
    post_save.connect(on_user_save, sender=EbekUser)
    post_delete.connect(on_user_delete, sender=EbekUser)
    post_save.connect(on_institute_save, sender=Institution)
    post_delete.connect(on_institution_delete, sender=Institution)
    post_save.connect(on_hospital_save, sender=Hospital)
    post_delete.connect(on_hospital_delete, sender=Hospital)
    post_save.connect(on_learner_save, sender=Learner)
    post_delete.connect(on_learner_delete, sender=Learner)
    post_save.connect(on_assessor_save, sender=Assessor)
    post_delete.connect(on_assessor_delete, sender=Assessor)
    post_save.connect(on_skillathon_save, sender=SkillathonEvent)
    post_delete.connect(on_skillathon_delete, sender=SkillathonEvent)
    post_save.connect(on_group_save, sender=Group)
    post_delete.connect(on_group_delete, sender=Group)
    print("All signals have been reconnected.")

@login_required
def learner_list_api(request):
    if not (request.user.has_all_permissions() or 'view_learners' in request.user.get_all_permissions()):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 10))
    search_query = request.GET.get('search', '').strip()
    
    print(f"DEBUG: Learner search query received: '{search_query}'")
    assigned_colleges = request.user.assigned_institutions.values_list('id', flat=True)
    assigned_hospitals = request.user.assigned_hospitals.values_list('id', flat=True)
    # Filtering logic (same as main view)
    if request.user.has_all_permissions():
        learners = Learner.objects.all().order_by('-created_at')
    else:
        learners = Learner.objects.filter(
            Q(college_id__in=assigned_colleges) | Q(hospital_id__in=assigned_hospitals)
        ).order_by('-created_at')
    
    # Apply search filter if search query exists
    if search_query:
        print(f"DEBUG: Applying learner search filter for: '{search_query}'")
        learners = learners.filter(
            Q(learner_user__full_name__icontains=search_query) |
            Q(learner_user__email__icontains=search_query) |
            Q(learner_user__phone_number__icontains=search_query) |
            Q(learner_type__icontains=search_query)
        )
    
    selected_institutions = request.GET.getlist('institution')
    selected_hospitals = request.GET.getlist('hospital')
    selected_learner_types = request.GET.getlist('learner_type')

    if selected_institutions:
        learners = learners.filter(college_id__in=selected_institutions)
    if selected_hospitals:
        learners = learners.filter(hospital_id__in=selected_hospitals)
    if selected_learner_types:
        learners = learners.filter(learner_type__in=selected_learner_types)

    # Get total count for pagination info
    total_count = learners.count()
    print(f"DEBUG: Total learners after filtering: {total_count}")
    
    # Pagination (offset/limit)
    learners = learners[offset:offset+limit]
    
    # Check if all data is loaded
    all_loaded = (offset + limit) >= total_count

    data = []
    for learner in learners:
        data.append({
            'id': learner.pk,
            'full_name': learner.learner_user.full_name,
            'email': learner.learner_user.email,
            'phone_number': learner.learner_user.phone_number,
            'learner_type': learner.get_learner_type_display(),
            'institution': getattr(learner.college, 'name', '-') if learner.learner_type == 'student' else getattr(learner.hospital, 'name', '-'),
            'skillathon': learner.skillathon_event.name if learner.skillathon_event else '-',
            'is_active': learner.is_active,
            'edit_url': reverse('learner_edit', args=[learner.pk]),
            'delete_url': reverse('learner_delete', args=[learner.pk]),
        })
    
    print(f"DEBUG: Returning {len(data)} learners")
    
    return JsonResponse({
        'learners': data,
        'total_count': total_count,
        'all_loaded': all_loaded,
        'search_query': search_query,
    })

@login_required
def sync_strength_counts(request):
    """API to sync total_strength and nurse_strength from Firebase before loading lists"""
    try:
        print("DEBUG: Starting strength count sync from Firebase")
        
        # Get all institutions and hospitals
        institutions = Institution.objects.all()
        hospitals = Hospital.objects.all()
        
        # Count students from Firebase for institutions
        for institution in institutions:
            try:
                # Query Firebase for students in this institution
                students = db.collection('Users').where('role', '==', 'student').where('institution', '==', str(institution.name)).stream()
                student_count = len(list(students))
                print(f"DEBUG: Student count for {institution.name}: {student_count}")
                
                # Update institution total_strength
                if institution.total_strength != student_count:
                    institution.total_strength = student_count
                    institution.save()
                    print(f"DEBUG: Updated institution {institution.name} total_strength to {student_count}")
                
            except Exception as e:
                print(f"DEBUG: Error syncing institution {institution.name}: {str(e)}")
        
        # Count nurses from Firebase for hospitals
        for hospital in hospitals:
            try:
                # Query Firebase for nurses in this hospital
                nurses = db.collection('Users').where('role', '==', 'nurse').where('hospital_id', '==', str(hospital.id)).stream()
                nurse_count = len(list(nurses))
                
                # Update hospital nurse_strength
                if hospital.nurse_strength != nurse_count:
                    hospital.nurse_strength = nurse_count
                    hospital.save()
                    print(f"DEBUG: Updated hospital {hospital.name} nurse_strength to {nurse_count}")
                
            except Exception as e:
                print(f"DEBUG: Error syncing hospital {hospital.name}: {str(e)}")
        
        print("DEBUG: Strength count sync completed successfully")
        return JsonResponse({'success': True, 'message': 'Strength counts synced successfully'})
        
    except Exception as e:
        print(f"DEBUG: Error in sync_strength_counts: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)